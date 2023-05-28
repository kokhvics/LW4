import unittest
from flask import Flask, request
from openpyxl import load_workbook
from app import soc_opros, app


class TestSocOpros(unittest.TestCase):
    def setUp(self):
        self.app = app.test_client()

    def test_write_to_excel(self):
        # Входные данные для теста
        name = "Test User"
        age = "25"
        favcolor = "Green"
        zodiac = "Taurus"
        email = "test@example.com"

        # Вызываем метод POST для добавления новых данных в Excel-файл
        response = self.app.post('/test', data=dict(
            name=name,
            age=age,
            favcolor=favcolor,
            zodiac=zodiac,
            email=email
        ))

        # Проверяем, что сервер вернул код успешного завершения операции (200)
        self.assertEqual(response.status_code, 200)

        # Проверяем, что данные были успешно добавлены в Excel-файл
        workbook = load_workbook("base.xlsx")
        sheet = workbook.active
        row_count = sheet.max_row
        self.assertEqual(sheet.cell(row=row_count, column=1).value, name)
        self.assertEqual(sheet.cell(row=row_count, column=2).value, age)
        self.assertEqual(sheet.cell(row=row_count, column=3).value, favcolor)
        self.assertEqual(sheet.cell(row=row_count, column=4).value, zodiac)
        self.assertEqual(sheet.cell(row=row_count, column=5).value, email)



if __name__ == '__main__':
    unittest.main()