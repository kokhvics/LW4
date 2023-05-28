from flask import Flask, render_template, url_for, request, redirect
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def glavnaya():
    return render_template("index.html")

@app.route('/test', methods=['post','get'])
def soc_opros():
    if request.method == 'POST':
        name = request.form.get('name')
        age = request.form.get('age')
        favcolor = request.form.get('favcolor')
        zodiac = request.form.get('zodiac')
        email = request.form.get('email')
        # Создадим новый файл Excel и добавим новый лист
        workbook = load_workbook("base.xlsx")
        sheet = workbook.active

        # Добавим заголовки для первой строки
        sheet['A1'] = 'Имя'
        sheet['B1'] = 'Возраст'
        sheet['C1'] = 'Цвет'
        sheet['D1'] = 'Знак зодиака'
        sheet['E1'] = 'E-mail'

        # Получим номер следующей строки, чтобы добавить новые данные
        row_number = sheet.max_row + 1

        # Добавим данные в новую строку
        sheet.cell(row=row_number, column=1, value=name)
        sheet.cell(row=row_number, column=2, value=age)
        sheet.cell(row=row_number, column=3, value=favcolor)
        sheet.cell(row=row_number, column=4, value=zodiac)
        sheet.cell(row=row_number, column=5, value=email)

        # Сохраним новый файл Excel
        workbook.save('base.xlsx')
        return render_template("submit.html", ans="Спасибо за заполнение анкеты!") #перенаправление на сайт submit
    # Вернем сообщение пользователю
    return render_template("soc_opros.html")

@app.route('/anketa', methods=['post','get'])
def quiz():
    if request.method == 'POST':
        option = request.form.get('options')

        if option == 'option1':
            return redirect('https://www.sunhome.ru/i/wallpapers/209/angry-birds.orig.jpg')
        elif option == 'option2':
            return redirect('https://2.bp.blogspot.com/-ZcmEBnO-JvU/WGQqpeweXpI/AAAAAAAAOqY/92A7lxdwF0AVpe8Y519q5DRDGDMQ-nH-ACLcB/s1600/ABT101.3.png')
        elif option == 'option3':
            return redirect('https://i.ytimg.com/vi/xxFVrzDVad4/maxresdefault.jpg')
        elif option == 'option4':
            return redirect('https://i.ytimg.com/vi/tRtxSmOnecU/maxresdefault.jpg')
        elif option == 'option5':
            return redirect('https://i.ytimg.com/vi/xxFVrzDVad4/maxresdefault.jpg')
        else:
            return redirect('https://example.com/default.jpg')
    return render_template("anketa.html")

@app.route('/about')
def about():
    return render_template("about.html")